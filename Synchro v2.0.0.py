#Synchro v2.0.0
import pandas as pd #for working with dataframes
import openpyxl as xl #for working with Excel files
from openpyxl.styles import Font, PatternFill
import shutil #for copying files
import math


COLUMN_ROOM = '[Common][Rooms][][]'
COLUMN_INSTANCE_NAME = '[Common][Name][][]'
COLUMN_TYPE_NAME = '[Identity Data][Type Name][][]'
COLUMN_AREA = '[Dimensions][Area][squareMeters][]'
COLUMN_DIMENSION = '[Dimensions][Length][millimeters][]'
COLUMN_UNICLASS_NUMBER = 'type[Data][Classification.Uniclass.Ss.Number][][]'
COLUMN_UNICLASS_DESCR = 'type[Data][Classification.Uniclass.Ss.Description][][]'
COLUMN_KEYNOTE = 'type[Constraints][Tag_Code_WB][][]'

SOURCE_FILE = ''
TARGET_FILE = ''
ORIGINAL_TAB = 'PBS Template'

FIELD_TO_CELL = {
    'Description': 'B6', # uniclass Ss description     # UNICLASS GROUP
    'Uniclass': 'B9', # uniclass Ss number      # UNICLASS GROUP
    'Name': 'A15', # instance name + type name
    'Qty': 'D15', # sum of areas or count of elements or sum or lengths
    'Unit': 'E15' # m2 or EA or mm
}



class ExcelHelper:
    '''Helper class for Excel operations'''
    def __init__(self):
        return

    def get_column_row(self, field, count = 0):
        # get column and row from cell name
        column = ''.join(filter(str.isalpha, FIELD_TO_CELL.get(field)))
        row = ''.join(filter(str.isdigit, FIELD_TO_CELL.get(field)))
        return column, int(row) + count

class UniClassGroup:
    '''Class for grouping elements in a room'''
    def __init__(self, df_group):
        print('initializing group')
        self.group = df_group
        self.code = None
        self.description = None
        self.populate_identity()
        print(f'group code = {self.code}')
        print(f'group description = {self.description}')
        self.elements = self.populate_elements()

    # method to populate identity information to a room group based on the tandem data in the room column
    # 'self' is the room group
    def populate_identity(self):
        # get the first row in the group
        first_row = self.group.iloc[0]
        # get the uniclass code from the first row
        self.code = first_row[COLUMN_UNICLASS_NUMBER]
        # get the uniclass description from the first row
        self.description = first_row[COLUMN_UNICLASS_DESCR]
        return

    def populate_elements(self):
        # group elements in group by type name
        df_elements = self.group.groupby(COLUMN_TYPE_NAME)
        print(f'got {len(df_elements)} elements in that uniclass group')

        groups = []
        # iterate over elements and create class
        for name, elements in df_elements:
            print(f'creating element {name}')
            elementG = ElementGroup(name, elements)
            groups.append(elementG)
        return groups

class ElementGroup:
    '''Class containing type group'''
    def __init__(self, name, df_group):
        print(f'initializing element {name}')
        self.group = df_group
        self.name = None
        self.qty = 0
        self.unit = None
        self.populate_identity()
        self.calculate_qty()
        
    def populate_identity(self):
        self.name = self.group.loc[self.group.index[0], COLUMN_INSTANCE_NAME] + ' : ' + self.group.loc[self.group.index[0], COLUMN_TYPE_NAME]

    def calculate_qty(self):
        # get sum of areas
        areas = self.group[COLUMN_AREA]

        # Check if all values in 'areas' are numeric
        all_numeric = all(isinstance(val, (int, float)) and not math.isnan(val) for val in areas)

        if all_numeric:
            self.qty = sum(areas)
            self.unit = 'm2'
        else:
            # check if all values in lengths are numeric
            lengths = self.group[COLUMN_DIMENSION]
            
            all_numeric = all(isinstance(val, (int, float)) and not math.isnan(val) for val in lengths)
            
            if all_numeric:
                self.qty = sum(lengths)
                self.unit = 'mm'
                
            else :
                self.qty = len(self.group)
                self.unit = 'EA'

        if self.qty == 0:
            self.qty = len(self.group)
            self.unit = 'EA'

        print(f'quantity for {self.name} = {self.qty} {self.unit}')

        return

class NewSheet:
    '''Class representing and creating a new sheet in an Excel file'''
    def __init__(self, wb, uniclass_group):
        print(f'initializing new sheet for room {uniclass_group.code}')
        self.uniclass_group = uniclass_group
        self.description = uniclass_group.description
        self.code = uniclass_group.code
        self.wb = wb
        self.sheet = self.create_tab(self.code)
        if self.sheet:
            self.populate_tab()
        else:
            print('could not create tab')
        return

    def create_tab(self, tab_name):
        try:
            print(f'creating tab {tab_name}')
            if tab_name in self.wb.sheetnames:
              existing_sheet = self.wb[tab_name]
              self.wb.remove(existing_sheet)
            #get source tab by name
            source = self.wb[ORIGINAL_TAB]
            # create the new tab
            new_sheet = self.wb.copy_worksheet(source)
            # name the new tab
            new_sheet.title = tab_name
            # Save the changes to the Excel file
            return new_sheet
        except Exception as e:
            print(f"An error occurred: {str(e)}")
            return None

    def populate_tab(self):
        xlh = ExcelHelper()
        count = 0
        print(f'populating tab {self.code}')

        column,row = xlh.get_column_row('Description')
        self.sheet[f'{column}{row}'] = self.uniclass_group.description
        print(f'wrote {self.uniclass_group.description} to {column}{row}')
        
        column,row = xlh.get_column_row('Uniclass')
        self.sheet[f'{column}{row}'] = self.uniclass_group.code
        print(f'wrote {self.uniclass_group.code} to {column}{row}')

        if(len(self.uniclass_group.elements) > 17):
            amount = len(self.uniclass_group.elements)-17
            column,row = xlh.get_column_row('Name')
            print(f'inserted {amount} rows at row {row}')

            for cell in self.sheet[row + 1]:
                    cell.font = Font(name='Arial', size=8)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    try:
                        if cell.is_merged:
                            self.sheet.unmerge_cells(cell.coordinate)
                    except Exception as e:
                        pass

            self.sheet.insert_rows(row+1, amount= amount)
            
            for i in range(amount):
                for cell in self.sheet[row + 1 + i]:
                    cell.font = Font(name='Arial', size=8)
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    try:
                        if cell.is_merged:
                            self.sheet.unmerge_cells(cell.coordinate)

                    except Exception as e:
                        pass

        for element in self.uniclass_group.elements:

            # get column and row from cell name
            column, row = xlh.get_column_row('Name', count)
            # write value to cell
            self.sheet[f'{column}{row}'] = element.name
            # get column and row from cell name
            column, row = xlh.get_column_row('Qty', count)
            # write value to cell
            self.sheet[f'{column}{row}'] = element.qty
            # get column and row from cell name
            column, row = xlh.get_column_row('Unit', count)
            # write value to cell
            self.sheet[f'{column}{row}'] = element.unit
            for cell in self.sheet[row]:
                try:
                    if cell.is_merged:
                        self.sheet.unmerge_cells(cell.coordinate)

                except Exception as e:
                    pass
                
            count += 1
        return


print('setup finished')

SOURCE_FILE = 'content/source.xlsx'
TARGET_FILE = 'content/dest.xlsx'



if SOURCE_FILE is None or SOURCE_FILE == '':
  print('SOURCE_FILE NOT UPLOADED, EXITING SCRIPT')
  exit()

if TARGET_FILE is None or TARGET_FILE == '':
  print('TARGET_FILE NOT UPLOADED, EXITING SCRIPT')
  exit()

if '.csv' in SOURCE_FILE:
  df = pd.read_csv(SOURCE_FILE)
else:
  df = pd.read_excel(SOURCE_FILE, engine='openpyxl')

#remove first row
df = df.iloc[1:]
# group by room
groups = df.groupby(COLUMN_UNICLASS_NUMBER)
uniclasses = []

print('got dataframe and groups')

# iterate over group and create RoomGroup class
# extraction of data and grouping by type happens inside the class.
# See RoomGroup init function
for name, group in groups:
    if name is not None and name != '' and name != ' ':
      print(f'creating uniclass group {name}')
      # create class from group
      uni = UniClassGroup(group)
      # append class to list of rooms
      uniclasses.append(uni)


for uni in uniclasses:
    print(f'uniclass code = {uni.code}')
    print(f'uniclass description = {uni.description}')
    print(f'uniclass elements = {len(uni.elements)}')



if TARGET_FILE is None or TARGET_FILE == '':
  print('TARGET_FILE NOT UPLOADED, EXITING SCRIPT')
  exit()

# load TARGET_FILE as workbook
wb = xl.load_workbook(TARGET_FILE)
# create new tab for each class
for uni in uniclasses:
    print(f'creating tab for {uni.code}')
    sheet = NewSheet(wb, uni)
    # populate tab with data

# save workbook
wb.save('merged.xlsx')

import os
for filename in os.listdir('content/'):
    if filename != 'merged.xlsx':
        file_path = os.path.join('content/', filename)
        #if os.path.isfile(file_path):
            #os.remove(file_path)

print('----------------------')
print('----------------------')
print('')
print('----------------------')
print('----------------------')
print(f'FILE SAVED AS : MERGED.XLSX')